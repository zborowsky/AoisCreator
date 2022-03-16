import datetime
import math
import re

import openpyxl
import os
from rich.progress import Progress
import logging
from rich.logging import RichHandler


class TimestampFileParser:
    def __init__(self, timestamp_dir, config_file):
        self.timestamp_dir = timestamp_dir
        self.time_column_name = "Time"
        self.distance_column_name = "Distance[km]"
        self.offset_column_name = "offset T"
        self.event_name = "event"
        self.config_file = config_file
        self.log = logging.getLogger("rich")

    def get_timestamps_files(self):
        return [self.timestamp_dir + "/" + file for file in os.listdir(self.timestamp_dir) if ".xlsx" in file]

    def get_column_sign(self, file_dir, worksheet):
        column_signs = {}
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == self.time_column_name:
                    column_signs[self.time_column_name] = cell.column
                elif cell.value == self.distance_column_name:
                    column_signs[self.distance_column_name] = cell.column
                elif cell.value == self.offset_column_name:
                    column_signs[self.offset_column_name] = cell.column
                elif cell.value == self.event_name:
                    column_signs[self.event_name] = cell.column

        if self.time_column_name not in column_signs:
            self.log.error("Time column not found in {file}, current column name is: {time_name}".format(file=file_dir, time_name=self.time_column_name))
            exit(-1)
        elif self.distance_column_name not in column_signs:
            self.log.error("Distance column name not found in {file}, searched for column with name: {distance_name}".format(file=file_dir, distance_name=self.distance_column_name))
            exit(-1)
        elif self.offset_column_name not in column_signs:
            self.log.error("Offset column name not found in {file}, searched for column with name: {offset_name}".format(file=file_dir, offset_name=self.offset_column_name))
            exit(-1)
        elif self.event_name not in column_signs:
            self.log.error("Event column name not found in {file}, searched for column with name: {event_name}".format(file=file_dir, event_name=self.event_name))
            exit(-1)

        self.log.info("Columns signs found in {file} successfully".format(file=file_dir))
        return column_signs

    def get_start_stop_offset(self, worksheet, column_signs):
        result = {}
        for col in worksheet.iter_cols(min_col=column_signs[self.event_name], max_col=column_signs[self.event_name]):
            for cell in col:
                if "start_offset" in result and "end_offset" in result:
                    break

                if cell.value == "start":
                    start_offset_value = worksheet.cell(cell.row, column_signs[self.offset_column_name]).value
                    result["start_offset_value"] = round(start_offset_value, 4)
                    result["start_offset_row"] = cell.row
                elif cell.value == "end":
                    end_offset_value = worksheet.cell(cell.row, column_signs[self.offset_column_name]).value
                    result["end_offset_value"] = round(end_offset_value, 4)
                    result["end_offset_row"] = cell.row

        return result

    @staticmethod
    def remove_extension(filename, extensions=".xlsx"):
        for extension in extensions:
            filename = filename.replace(extension, "")
        return filename

    def show_progress(self, current_number, all_to_proceed_number):
        os.system("cls")
        if all_to_proceed_number != 1:
            self.log.info("Found ", all_to_proceed_number, "timestamps files")
            self.log.info("Timestamp data processing progress: " + str(
                round(current_number / (all_to_proceed_number - 1) * 100, 2)) + "%")

    def get_aois_data(self):
        with Progress() as progress:
            user_dicts = {}
            aois_timing = {}
            timestamp_files = self.get_timestamps_files()
            task1 = progress.add_task("[green]Updating EventConfigs.json", total=len(timestamp_files))
            while not progress.finished:
                for file in timestamp_files:
                    excel_worksheet = openpyxl.load_workbook(file, data_only=True).active
                    participant_id = re.search(r"\((.*)\)", file.split("/")[-1]).group(1)
                    column_signs = self.get_column_sign(file, excel_worksheet)
                    user_dicts[participant_id] = self.get_start_stop_offset(excel_worksheet, column_signs)
                    aois_timing_data = self.get_data_from_mileage(excel_worksheet, column_signs, user_dicts[participant_id])
                    aois_timing[participant_id] = self.allign_aoi_events_with_offset(aois_timing_data, user_dicts[participant_id])
                    progress.update(task1, advance=1)
            return user_dicts, aois_timing, column_signs

    def update_millage_config_based_on_time(self, user_dicts, duration_data, column_signs):

        timestamp_files = self.get_timestamps_files()
        participant_id_regex = re.compile(r"\((.*)\)")
        sample_video_participant_id = re.search(participant_id_regex, duration_data["SampleVideoName"]).group(1)
        for timestamp_file in timestamp_files:
            candidate_participant_id = re.search(participant_id_regex, timestamp_file.split("/")[-1]).group(1)
            if candidate_participant_id == sample_video_participant_id:
                excel_worksheet = openpyxl.load_workbook(timestamp_file, data_only=True).active
                return self.get_millage_from_time(excel_worksheet, user_dicts[sample_video_participant_id], duration_data, column_signs)

        self.log.error("No suitable timestamp found")
        exit(-1)

    @staticmethod
    def allign_aoi_events_with_offset(aois_timings, offset_dict):
        adjusted_timing = {}
        for [aoi_timing_name, aois_timing_list] in aois_timings.items():
            adjusted_aois_timing_list = []
            for aoi_timing_event in aois_timing_list:
                adjusted_aois_timing_list.append(aoi_timing_event + offset_dict["start_offset_value"])
            adjusted_timing[aoi_timing_name] = adjusted_aois_timing_list
        return adjusted_timing

    def get_data_from_mileage(self, excel_worksheet, column_signs, start_offset_coordinate):
        result = {}
        start_row_number = start_offset_coordinate["start_offset_row"]
        time_column_sign = column_signs[self.time_column_name]
        start_h, start_m, start_s, start_mm = excel_worksheet.cell(start_row_number, time_column_sign).value.split(":")
        start_time = datetime.timedelta(hours=int(start_h), minutes=int(start_m), seconds=int(start_s))
        for [event_name, event_data] in self.config_file.items():
            event_timings = []
            for occurrence_millage in event_data["VideoOccurrences"]:
                for column in excel_worksheet.iter_cols(min_col=column_signs[self.distance_column_name],
                                                        max_col=column_signs[self.distance_column_name], min_row=2):
                    for cell in column:
                        if float(occurrence_millage) <= float(cell.value):
                            event_h, event_m, event_s, event_mm = excel_worksheet.cell(cell.row, time_column_sign).value.split(":")
                            event_time = datetime.timedelta(hours=int(event_h), minutes=int(event_m), seconds=int(event_s))
                            event_timings.append((event_time - start_time).seconds)
                            break
            result[event_name] = event_timings

        return result

    @staticmethod
    def from_seconds_to_time(float_time):
        hours = 0
        minutes = 0

        while float_time >= 60:
            float_time -= 60
            minutes += 1
            if minutes >= 60:
                minutes = 0
                hours += 1

        result = "{h}:{m}:{s}".format(h=hours, m=minutes, s=math.floor(float_time))
        return result

    def get_millage_from_time(self, excel_worksheet, start_stop_offset, duration_config, column_signs):
        video_offset_h, video_offset_m, video_offset_s = self.from_seconds_to_time(start_stop_offset["start_offset_value"]).split(":")
        video_offset = datetime.timedelta(hours=int(video_offset_h), minutes=int(video_offset_m), seconds=int(video_offset_s))

        result = {}
        start_row_number = start_stop_offset["start_offset_row"]
        time_column_sign = column_signs[self.time_column_name]
        start_h, start_m, start_s, start_mm = excel_worksheet.cell(start_row_number, time_column_sign).value.split(":")
        start_time = datetime.timedelta(hours=int(start_h), minutes=int(start_m), seconds=int(start_s))

        for [aoiName, occurrencesList] in duration_config.items():
            if aoiName != "SampleVideoName":
                aoi_occurrence_millage = []
                for occurrence_time in occurrencesList:
                    occurrence_h, occurrence_m, occurrence_s = occurrence_time.split(":")
                    occurrence_time_delta = datetime.timedelta(hours=int(occurrence_h), minutes=int(occurrence_m), seconds=int(occurrence_s))
                    occurrence_with_offset_adjusted = (occurrence_time_delta - video_offset).seconds
                    for column in excel_worksheet.iter_cols(min_col=column_signs[self.time_column_name],
                                                            max_col=column_signs[self.time_column_name], min_row=start_stop_offset["start_offset_row"]):
                        for cell in column:
                            event_h, event_m, event_s, event_mm = excel_worksheet.cell(cell.row, column_signs[self.time_column_name]).value.split(":")
                            event_time_delta = datetime.timedelta(hours=int(event_h), minutes=int(event_m), seconds=int(event_s))
                            event_time_delta_with_offset_adjusted = (event_time_delta - start_time).seconds
                            if occurrence_with_offset_adjusted <= event_time_delta_with_offset_adjusted:
                                aoi_occurrence_millage.append(float(excel_worksheet.cell(cell.row, column_signs[self.distance_column_name]).value))
                                break

                result[aoiName] = aoi_occurrence_millage
        return result