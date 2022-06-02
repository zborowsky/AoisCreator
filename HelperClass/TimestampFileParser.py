import datetime
import re

import openpyxl
import os

import pandas
from rich.progress import Progress

from HelperClass.HelperMethods import get_calibration_time, check_if_exported_data_exists, get_particpant_id
from HelperClass.RichHelper import console, error, warning, info


class TimestampFileParser:
    def __init__(self, timestamp_dir, config_file):
        self.timestamp_dir = timestamp_dir
        self.time_column_name = "Time"
        self.distance_column_name = "Distance[km]"
        self.offset_column_name = "offset T"
        self.config_file = config_file

    def get_timestamps_files(self):
        try:
            return [self.timestamp_dir + "/" + file for file in os.listdir(self.timestamp_dir) if ".xlsx" in file]
        except Exception as err:
            console.print("ERR {err}".format(err=err), style=error)
            exit(-1)

    def get_column_sign(self, file_dir, worksheet):
        column_signs = {}
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == self.time_column_name:
                    column_signs[self.time_column_name] = cell.column
                elif cell.value == self.distance_column_name:
                    column_signs[self.distance_column_name] = cell.column
                elif cell.value == self.offset_column_name:
                    column_signs[self.offset_column_name] = cell.column

        necessary_columns = {
            self.time_column_name, self.distance_column_name, self.offset_column_name
        }

        if not all(col in column_signs for col in necessary_columns):
            console.print("\nERR Some necessary column not found in {file}".format(file=file_dir), style=error)
            console.print("ERR Necessary columns are:", style=error)
            for col in necessary_columns:
                console.print("{col}".format(col=col), style=warning)
            return None
        else:
            return column_signs


    @staticmethod
    def check_if_timestamp_in_timestamp_list(timestamp_files, selected_timestamp_name):
        result = [timestamp for timestamp in timestamp_files if selected_timestamp_name in timestamp]
        result_len = len(result)

        if result_len == 0:
            console.print("ERR SampleVideoName timestamp could not be found in provided dir", style=error)
            exit(-1)
        elif result_len > 1:
            console.print("ERR More than one SampleVideoName timestamp was found", style=error)
            exit(-1)
        else:
            return result

    def get_aois_data(self, exported_data_dir, update_config_timestamp=""):
        with Progress() as progress:
            aois_timing = dict()
            column_signs = dict()
            timestamp_files = self.get_timestamps_files()

            if update_config_timestamp:
                timestamp_files = self.check_if_timestamp_in_timestamp_list(timestamp_files, update_config_timestamp)
                selected_id = get_particpant_id(update_config_timestamp)
                if not check_if_exported_data_exists(exported_data_dir, selected_id):
                    exit(-1)

            task1 = progress.add_task("[green]Processing timestamp data", total=len(timestamp_files))
            while not progress.finished:
                for file in timestamp_files:
                    participant_id = re.search(r"\((.*)\)", file.split("/")[-1])
                    if participant_id:
                        participant_id = participant_id.group(1)

                        excel_worksheet = openpyxl.load_workbook(file, data_only=True).active
                        column_signs = self.get_column_sign(file, excel_worksheet)
                        if not column_signs:
                            console.print("ERR {file} will be omitted".format(file=file), style=error)
                            continue
                        else:
                            aois_timing_data = self.get_data_from_mileage(exported_data_dir, excel_worksheet,
                                                                          participant_id, column_signs)
                            aois_timing[participant_id] = aois_timing_data
                    else:
                        console.print("\nERR Timestamp filename should consist participant id placed in () for example xxx(u01).xlsx", style=error)
                    progress.update(task1, advance=1)
        return aois_timing, column_signs

    def update_millage_config_based_on_time(self, export_data_dir, duration_data, column_signs):
        timestamp_files = self.get_timestamps_files()
        participant_id_regex = re.compile(r"\((.*)\)")
        sample_video_participant_id = re.search(participant_id_regex, duration_data["SampleVideoName"]).group(1)
        for timestamp_file in timestamp_files:
            candidate_participant_id = re.search(participant_id_regex, timestamp_file.split("/")[-1]).group(1)
            if candidate_participant_id == sample_video_participant_id:
                excel_worksheet = openpyxl.load_workbook(timestamp_file, data_only=True).active
                return self.get_millage_from_time(export_data_dir, sample_video_participant_id, excel_worksheet, duration_data, column_signs)

        console.print("ERR No suitable timestamp found", style=error)
        exit(-1)

    def get_data_from_mileage(self, exported_data_dir, excel_worksheet, participant_id, column_signs):
        calibration_time = get_calibration_time(exported_data_dir, participant_id)
        result = {}
        if calibration_time is not None:
            for [event_name, event_data] in self.config_file.items():
                event_timings = []
                for occurrence_entry in event_data["VideoOccurrences"]:
                    for column in excel_worksheet.iter_cols(min_col=column_signs[self.distance_column_name],
                                                            max_col=column_signs[self.distance_column_name], min_row=2):
                        for cell in column:
                            if float(occurrence_entry["Millage"]) <= float(cell.value):
                                game_offset_time = round(float(excel_worksheet.cell(cell.row, column_signs[self.offset_column_name]).value), 3)
                                game_offset_time_list = str(game_offset_time).split(".")
                                if len(game_offset_time_list) == 2:
                                    game_offset_time_s = game_offset_time_list[0]
                                    game_offset_time_ms = game_offset_time_list[1]
                                else:
                                    game_offset_time_s = game_offset_time_list[0]
                                    game_offset_time_ms = "0"

                                game_offset_time_delta = datetime.timedelta(seconds=int(game_offset_time_s), milliseconds=int(game_offset_time_ms))
                                event_timings.append((game_offset_time_delta - calibration_time).total_seconds())
                                break
                result[event_name] = event_timings

        return result

    def get_millage_from_time(self, export_data_dir, sample_participant_id, excel_worksheet, duration_config, column_signs):
        game_start_offset = excel_worksheet.cell(2, column_signs[self.offset_column_name]).value
        game_start_offset_s, game_start_offset_micro_s = str(round(game_start_offset, 6)).split(".")

        game_start_time = excel_worksheet.cell(2, column_signs[self.time_column_name]).value
        game_start_time_h, game_start_time_m, game_start_time_s, game_start_time_milli_s = str(game_start_time).split(":")

        game_start_time_delta = datetime.timedelta(hours=int(game_start_time_h), minutes=int(game_start_time_m),
                                                   seconds=int(game_start_time_s), milliseconds=int(game_start_time_milli_s))

        game_start_offset_delta = datetime.timedelta(seconds=int(game_start_offset_s), microseconds=int(game_start_offset_micro_s))

        calibration_time = get_calibration_time(export_data_dir, sample_participant_id)
        offset_game_recording_start = game_start_offset_delta - calibration_time

        result = {}
        for [aoiName, occurrencesList] in duration_config.items():
            if aoiName != "SampleVideoName":
                aoi_occurrence_millage = []
                for occurrence_entry in occurrencesList:
                    occurrence_h, occurrence_m, occurrence_s, occurrence_mic_s = occurrence_entry["Time"].split(":")
                    occurrence_time_delta = datetime.timedelta(hours=int(occurrence_h), minutes=int(occurrence_m), seconds=int(occurrence_s), microseconds=int(occurrence_mic_s))
                    occurrence_time_delta_to_game_start = occurrence_time_delta - offset_game_recording_start
                    game_millage_time = game_start_time_delta + occurrence_time_delta_to_game_start

                    for column in excel_worksheet.iter_cols(min_col=column_signs[self.time_column_name],
                                                            max_col=column_signs[self.time_column_name], min_row=2):
                        for cell in column:
                            event_h, event_m, event_s, event_mm = excel_worksheet.cell(cell.row, column_signs[self.time_column_name]).value.split(":")
                            event_time_delta = datetime.timedelta(hours=int(event_h), minutes=int(event_m), seconds=int(event_s))
                            if game_millage_time <= event_time_delta:
                                aoi_occurrence_millage.append({
                                    "AoiState": occurrence_entry["Status"],
                                    "Millage": float(excel_worksheet.cell(cell.row, column_signs[self.distance_column_name]).value)
                                })
                                break

                result[aoiName] = aoi_occurrence_millage
        return result
